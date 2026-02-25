
#!/usr/bin/env python3
"""
Universal HVAC ingestion + enrichment tool (site-specific extractors + manufacturer catalog + domain learning).

Key upgrades vs prior version:
- Loads a manufacturer catalog CSV: brand,domains (comma-separated).
- Works even when domains are blank by doing broad search first.
- Optional domain learning: stores the best hit's host per brand to learned_domains.json so future runs bias toward it.
  This is how you scale to "every manufacturer" without hand-curating 400 domains.

Dependencies:
  pip install openpyxl requests
Recommended:
  pip install beautifulsoup4 lxml
Optional (PDF):
  pip install pdfplumber
"""

from __future__ import annotations

import argparse
import dataclasses
import hashlib
import json
import random
import re
import fnmatch
import sqlite3
import sys
import time
import urllib.parse
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple

import requests
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

try:
    from bs4 import BeautifulSoup  # type: ignore
except Exception:
    BeautifulSoup = None  # type: ignore

try:
    import pdfplumber  # type: ignore
except Exception:
    pdfplumber = None  # type: ignore


# ----------------------------
# Utilities
# ----------------------------

def norm_model(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s).strip()).upper()

def norm_brand(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip()).upper()

def safe_sleep(base: float = 0.6, jitter: float = 0.4) -> None:
    time.sleep(max(0.0, base + random.random() * jitter))

def sha1(s: str) -> str:
    return hashlib.sha1(s.encode("utf-8", errors="ignore")).hexdigest()

def load_rules_json(path: str):
    try:
        if not path:
            return []
        p = Path(path)
        if not p.exists():
            return []
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return []

def first_rule_match(rules, brand: str, model: str):
    b = (brand or "").strip().upper()
    m = (model or "").strip().upper()
    for r in rules or []:
        if (r.get("brand","").strip().upper() not in ("", b)):
            continue
        pat = (r.get("pattern") or "").strip().upper()
        if not pat:
            continue
        if fnmatch.fnmatch(m, pat):
            return r
    return None

def clean_text(s: str) -> str:
    s = s or ""
    s = re.sub(r"<[^>]+>", " ", s)
    s = s.replace("&amp;", "&").replace("&nbsp;", " ").replace("&#39;", "'").replace("&quot;", '"')
    s = re.sub(r"\s+", " ", s).strip()
    return s

def first_nonempty(*vals: str) -> str:
    for v in vals:
        if v and str(v).strip():
            return str(v).strip()
    return ""

def truncate(s: str, n: int = 1800) -> str:
    s = s or ""
    return s if len(s) <= n else (s[:n].rstrip() + "…")

def host_of(url: str) -> str:
    try:
        return urllib.parse.urlparse(url).netloc.lower()
    except Exception:
        return ""


# ----------------------------
# Cache (SQLite)
# ----------------------------

class SQLiteCache:
    def __init__(self, path: str):
        self.path = path
        self.conn = sqlite3.connect(path)
        self.conn.execute(
            """
            CREATE TABLE IF NOT EXISTS cache (
              k TEXT PRIMARY KEY,
              v TEXT NOT NULL,
              ts INTEGER NOT NULL
            )
            """
        )
        self.conn.commit()

    def get(self, k: str) -> Optional[str]:
        cur = self.conn.execute("SELECT v FROM cache WHERE k=?", (k,))
        row = cur.fetchone()
        return row[0] if row else None

    def set(self, k: str, v: str) -> None:
        self.conn.execute(
            "INSERT OR REPLACE INTO cache(k,v,ts) VALUES(?,?,strftime('%s','now'))",
            (k, v),
        )
        self.conn.commit()

    def close(self):
        self.conn.close()


# ----------------------------
# Manufacturer catalog + learning
# ----------------------------

@dataclass
class ManufacturerProfile:
    name: str
    domains: List[str]

class ManufacturerCatalog:
    def __init__(self, csv_path: Optional[str] = None, learned_path: str = "learned_domains.json"):
        self.csv_path = csv_path
        self.learned_path = learned_path
        self.map: Dict[str, ManufacturerProfile] = {}
        self.learned: Dict[str, List[str]] = {}
        self._load()

    def _load(self):
        # learned domains
        try:
            with open(self.learned_path, "r", encoding="utf-8") as f:
                self.learned = json.load(f)
        except Exception:
            self.learned = {}

        if not self.csv_path:
            return
        try:
            with open(self.csv_path, "r", encoding="utf-8") as f:
                lines = f.read().splitlines()
        except Exception:
            return

        # very small CSV parser (brand,domains)
        header = True
        for line in lines:
            if header:
                header = False
                continue
            if not line.strip():
                continue
            parts = [p.strip() for p in line.split(",", 1)]
            brand = parts[0] if parts else ""
            domains = []
            if len(parts) > 1 and parts[1]:
                domains = [d.strip().lower() for d in parts[1].split(",") if d.strip()]
            if brand:
                self.map[norm_brand(brand)] = ManufacturerProfile(brand.strip(), domains)

    def get(self, brand: str) -> ManufacturerProfile:
        b = norm_brand(brand)
        prof = self.map.get(b, ManufacturerProfile(brand.strip() or "Unknown", []))
        learned = self.learned.get(b, [])
        # learned first (usually the most accurate for your data)
        domains = [d.lower() for d in learned if d] + [d.lower() for d in prof.domains if d]
        # de-dupe
        seen = set()
        uniq = []
        for d in domains:
            if d in seen:
                continue
            seen.add(d)
            uniq.append(d)
        return ManufacturerProfile(prof.name, uniq)

    def learn(self, brand: str, url: str) -> None:
        b = norm_brand(brand)
        h = host_of(url)
        if not h:
            return
        # keep apex-ish hosts only
        h = h.replace("www.", "")
        lst = self.learned.get(b, [])
        if h not in lst:
            lst = [h] + lst
            lst = lst[:8]
            self.learned[b] = lst
            try:
                with open(self.learned_path, "w", encoding="utf-8") as f:
                    json.dump(self.learned, f, indent=2)
            except Exception:
                pass


# ----------------------------
# Search (no API key)
# ----------------------------

@dataclass
class SearchResult:
    url: str
    title: str = ""

class SearchProvider:
    def search(self, query: str, *, max_results: int = 6) -> List[SearchResult]:
        raise NotImplementedError

class DuckDuckGoHTMLSearch(SearchProvider):
    ENDPOINT = "https://duckduckgo.com/html/"

    def __init__(self, session: requests.Session, cache: SQLiteCache, rate_limit: float = 1.0):
        self.s = session
        self.cache = cache
        self.rate_limit = rate_limit

    def search(self, query: str, *, max_results: int = 6) -> List[SearchResult]:
        ck = "ddg:" + sha1(query)
        cached = self.cache.get(ck)
        if cached:
            return [SearchResult(**r) for r in json.loads(cached)]

        safe_sleep(self.rate_limit, 0.5)
        resp = self.s.post(self.ENDPOINT, data={"q": query}, timeout=25)
        if resp.status_code != 200:
            self.cache.set(ck, "[]")
            return []

        html = resp.text or ""
        results: List[SearchResult] = []
        for m in re.finditer(r'class="result__a"\s+href="([^"]+)"[^>]*>(.*?)</a>', html, re.I | re.S):
            href = urllib.parse.unquote(m.group(1))
            title = clean_text(m.group(2))
            results.append(SearchResult(url=href, title=title))
            if len(results) >= max_results:
                break

        self.cache.set(ck, json.dumps([dataclasses.asdict(r) for r in results]))
        return results

class DomainPreferredSearch(SearchProvider):
    def __init__(self, base: SearchProvider, domains: List[str]):
        self.base = base
        self.domains = [d.lower().strip() for d in domains if d and d.strip()]

    def search(self, query: str, *, max_results: int = 6) -> List[SearchResult]:
        out: List[SearchResult] = []
        for d in self.domains:
            out.extend(self.base.search(f'site:{d} "{query}"', max_results=max_results))
            if len(out) >= max_results:
                return out[:max_results]
        out.extend(self.base.search(f'"{query}"', max_results=max_results))
        return out[:max_results]


# ----------------------------
# Fetcher
# ----------------------------

class Fetcher:
    def __init__(self, session: requests.Session, cache: SQLiteCache, *, rate_limit: float = 0.8):
        self.s = session
        self.cache = cache
        self.rate_limit = rate_limit

    def get(self, url: str) -> Tuple[str, str]:
        ck = "fetch:" + sha1(url)
        cached = self.cache.get(ck)
        if cached is not None:
            d = json.loads(cached)
            return d.get("ct",""), d.get("payload","")

        safe_sleep(self.rate_limit, 0.4)
        try:
            resp = self.s.get(url, timeout=35, allow_redirects=True, stream=True)
        except Exception:
            self.cache.set(ck, json.dumps({"ct":"", "payload":""}))
            return "", ""

        ct = (resp.headers.get("Content-Type") or "").lower()

        if "application/pdf" in ct or url.lower().endswith(".pdf"):
            data = resp.content
            tmp = f".cache_pdf_{sha1(url)}.pdf"
            try:
                with open(tmp, "wb") as f:
                    f.write(data)
                payload = "FILE:" + tmp
            except Exception:
                payload = ""
            self.cache.set(ck, json.dumps({"ct":ct, "payload":payload}))
            return ct, payload

        try:
            text = resp.text or ""
        except Exception:
            text = resp.content.decode("utf-8", errors="ignore")
        self.cache.set(ck, json.dumps({"ct":ct, "payload":text}))
        return ct, text


# ----------------------------
# Extractors (plug-ins)
# ----------------------------

class BaseExtractor:
    def matches(self, url: str, brand: str) -> bool:
        return True

    def extract(self, *, url: str, model: str, brand: str, content_type: str, payload: str) -> Tuple[str, str]:
        raise NotImplementedError

def _jsonld_product(html: str) -> Tuple[str, str]:
    name = ""
    desc = ""
    for m in re.finditer(r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>', html, re.I | re.S):
        raw = m.group(1).strip()
        raw = re.sub(r"^\s*<!--|-->\s*$", "", raw)
        try:
            data = json.loads(raw)
        except Exception:
            continue
        items = data if isinstance(data, list) else [data]
        for it in items:
            if not isinstance(it, dict):
                continue
            t = it.get("@type") or ""
            if isinstance(t, list):
                t = " ".join(map(str, t))
            if "product" in str(t).lower():
                name = first_nonempty(name, it.get("name",""))
                desc = first_nonempty(desc, it.get("description",""))
        if name or desc:
            break
    return clean_text(name), clean_text(desc)

def _meta_title_desc(html: str) -> Tuple[str, str]:
    og_title = ""
    meta_desc = ""
    for m in re.finditer(r"<meta\s+[^>]*>", html, re.I):
        tag = m.group(0)
        if re.search(r'property=["\']og:title["\']', tag, re.I):
            cm = re.search(r'content=["\']([^"\']+)["\']', tag, re.I)
            if cm:
                og_title = cm.group(1)
        if re.search(r'name=["\']description["\']', tag, re.I):
            cm = re.search(r'content=["\']([^"\']+)["\']', tag, re.I)
            if cm:
                meta_desc = cm.group(1)
    title = ""
    mt = re.search(r"<title[^>]*>(.*?)</title>", html, re.I | re.S)
    if mt:
        title = mt.group(1)
    h1 = ""
    mh = re.search(r"<h1[^>]*>(.*?)</h1>", html, re.I | re.S)
    if mh:
        h1 = mh.group(1)
    return clean_text(first_nonempty(og_title, h1, title)), clean_text(meta_desc)

def _soup_blocks(html: str, model: str) -> str:
    if not BeautifulSoup:
        return ""
    soup = BeautifulSoup(html, "lxml")  # type: ignore
    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()
    for sel in ["header", "footer", "nav"]:
        for t in soup.select(sel):
            t.decompose()

    keywords = ["overview","features","highlights","specifications","specs","product details","details","warranty","performance","efficiency"]
    chunks: List[str] = []

    for h in soup.find_all(["h1","h2","h3"]):
        htxt = clean_text(h.get_text(" ", strip=True)).lower()
        if any(k in htxt for k in keywords):
            cur = h
            acc = []
            steps = 0
            while cur and steps < 25:
                cur = cur.find_next_sibling()
                if cur is None:
                    break
                if getattr(cur, "name", "") in ["h1","h2","h3"]:
                    break
                t = clean_text(cur.get_text(" ", strip=True))
                if t and len(t) > 20:
                    acc.append(t)
                steps += 1
            if acc:
                chunks.append(clean_text(h.get_text(" ", strip=True) + ": " + " ".join(acc)))

    if model:
        patt = re.compile(re.escape(model), re.I)
        for p in soup.find_all(["p","li","div","span"]):
            t = clean_text(p.get_text(" ", strip=True))
            if t and patt.search(t) and len(t) > 30:
                chunks.append(t)

    # de-dupe
    seen = set()
    uniq = []
    for t in chunks:
        k = sha1(t.lower())
        if k in seen:
            continue
        seen.add(k)
        uniq.append(t)

    return truncate("\n".join(uniq), 2200)

class GenericExtractor(BaseExtractor):
    def extract(self, *, url: str, model: str, brand: str, content_type: str, payload: str) -> Tuple[str, str]:
        html = payload or ""
        jn, jd = _jsonld_product(html)
        mn, md = _meta_title_desc(html)
        body = _soup_blocks(html, model)

        name = first_nonempty(jn, mn)
        desc_parts = [x for x in [jd, md, body] if x]
        desc = "\n".join(desc_parts)
        return clean_text(name), truncate(desc, 2400)

class BryantExtractor(BaseExtractor):
    def matches(self, url: str, brand: str) -> bool:
        return "bryant.com" in host_of(url) or norm_brand(brand) == "BRYANT"

    def extract(self, *, url: str, model: str, brand: str, content_type: str, payload: str) -> Tuple[str, str]:
        name, desc = GenericExtractor().extract(url=url, model=model, brand=brand, content_type=content_type, payload=payload)
        if desc and model and model not in desc:
            desc += "\n" + f"Model: {model}"
        return name, desc

class CarrierExtractor(BaseExtractor):
    def matches(self, url: str, brand: str) -> bool:
        return "carrier.com" in host_of(url) or norm_brand(brand) == "CARRIER"

    def extract(self, *, url: str, model: str, brand: str, content_type: str, payload: str) -> Tuple[str, str]:
        name, desc = GenericExtractor().extract(url=url, model=model, brand=brand, content_type=content_type, payload=payload)
        if desc and model and model not in desc:
            desc += "\n" + f"Model: {model}"
        return name, desc



class DayNightExtractor(BaseExtractor):
    """Day & Night official site extractor (Carrier family brand)."""
    def matches(self, url: str, brand: str) -> bool:
        h = host_of(url)
        return ("dayandnightcomfort.com" in h) or (norm_brand(brand) in {"DAY & NIGHT", "DAY AND NIGHT", "DAY&NIGHT", "DAYNIGHT"})

    def extract(self, *, url: str, model: str, brand: str, content_type: str, payload: str) -> Tuple[str, str]:
        # Day & Night pages often behave like other Carrier-family AEM sites; generic extraction works fine.
        name, desc = GenericExtractor().extract(url=url, model=model, brand=brand, content_type=content_type, payload=payload)
        if desc and model and model not in desc:
            desc += "\n" + f"Model: {model}"
        return name, desc


class MitsubishiExtractor(BaseExtractor):
    """Mitsubishi Electric (M-Series / City Multi / etc.) official site extractor."""
    def matches(self, url: str, brand: str) -> bool:
        h = host_of(url)
        return ("mitsubishicomfort.com" in h) or (norm_brand(brand) in {"MITSUBISHI", "MITSUBISHI ELECTRIC"})

    def extract(self, *, url: str, model: str, brand: str, content_type: str, payload: str) -> Tuple[str, str]:
        # MitsubishiComfort pages usually include decent headings + sometimes JSON-LD.
        name, desc = GenericExtractor().extract(url=url, model=model, brand=brand, content_type=content_type, payload=payload)
        if desc and model and model not in desc:
            desc += "\n" + f"Model: {model}"
        return name, desc

class PDFExtractor(BaseExtractor):
    def matches(self, url: str, brand: str) -> bool:
        return url.lower().endswith(".pdf")

    def extract(self, *, url: str, model: str, brand: str, content_type: str, payload: str) -> Tuple[str, str]:
        if not payload.startswith("FILE:"):
            return "", ""
        path = payload.split("FILE:", 1)[1].strip()
        if not path:
            return "", ""
        if not pdfplumber:
            return "", f"Spec sheet PDF available.\nModel: {model}\nSource: {url}"
        try:
            text_parts: List[str] = []
            with pdfplumber.open(path) as pdf:  # type: ignore
                for page in pdf.pages[:3]:
                    t = page.extract_text() or ""
                    t = re.sub(r"\s+", " ", t).strip()
                    if t:
                        text_parts.append(t)
            full = " ".join(text_parts)
            first_sentence = full.split(". ")[0][:180]
            name = clean_text(first_sentence)
            desc = truncate(clean_text(full), 2400)
            return name, desc
        except Exception:
            return "", f"Spec sheet PDF available.\nModel: {model}\nSource: {url}"

EXTRACTORS: List[BaseExtractor] = [
    PDFExtractor(),
    BryantExtractor(),
    CarrierExtractor(),
    DayNightExtractor(),
    MitsubishiExtractor(),
    GenericExtractor(),
]


# ----------------------------
# Folder / warranty conventions (same as earlier)
# ----------------------------

TON3 = {"018":"1.5 Ton","024":"2 Ton","030":"2.5 Ton","036":"3 Ton","042":"3.5 Ton","048":"4 Ton","060":"5 Ton"}
TON2 = {"18":"1.5 Ton","24":"2 Ton","30":"2.5 Ton","36":"3 Ton","42":"3.5 Ton","48":"4 Ton","60":"5 Ton"}
FURN_CAB = {"14":'14.5"', "17":'17.5"', "21":'21"', "24":'24.5"'}

def find_ton(m: str) -> str:
    m3 = re.search(r"(018|024|030|036|042|048|060)", m)
    if m3:
        return TON3.get(m3.group(1), "")
    m2 = re.search(r"(18|24|30|36|42|48|60)", m)
    if m2:
        return TON2.get(m2.group(1), "")
    return ""

def find_btu_code(m: str) -> str:
    mm = re.search(r"(026|040|045|060|070|080|090|100|110|120|135|140|155)", m)
    return mm.group(1) if mm else ""

def warranty_block(brand: str) -> str:
    b = (brand or "").strip()
    return "\n".join([
        "WARRANTY",
        "• 10-year parts limited warranty with timely registration",
        "• If not registered within 90 days of installation, parts warranty defaults to 5 years",
        f"• See {b} limited warranty certificate for full terms, exclusions, and requirements",
    ])

@dataclass
class ParsedFolders:
    folder1: str
    folder2: str
    folder3: str
    folder4: str
    folder5: str

def parse_folders(brand: str, model: str) -> ParsedFolders:
    m = model
    f1 = "Equipment"
    f2 = brand
    folder3 = "Uncategorized"
    folder4 = ""
    folder5 = ""

    if re.match(r"^(987|986|927|926|916|880|820|800)", m):
        folder3 = "Gas Furnace"
        btu = find_btu_code(m)
        folder4 = f"{btu}k BTU" if btu else ""
        cm = re.search(r"[CVM](14|17|21|24)$", m)
        if cm:
            cab = FURN_CAB.get(cm.group(1), "")
            folder5 = f'{cab} Cabinet' if cab else ""
    elif re.match(r"^(191VAN|148TAN|146SAN|134SAN)", m):
        folder3 = "Air Conditioner"
        folder4 = find_ton(m)
    elif re.match(r"^(293VAN|290VAN|248TAN|246SAN|235SAN)", m):
        folder3 = "Heat Pump"
        folder4 = find_ton(m)
    elif re.match(r"^(FE5B|FT5A|FJ5A)", m):
        folder3 = "Fan Coil"
        folder4 = find_ton(m)
    elif re.match(r"^(EHK|CEHK|KSAC|KSAI|KFAD)", m):
        folder3 = "Heat Strip"
        kw = re.search(r"(05|08|10|15|20|25)", m)
        folder4 = f"{int(kw.group(1))} kW" if kw else ""
    elif re.match(r"^(CVAVA|CVAMA|CAAMP|CSAHP)", m):
        folder3 = "Evaporator Coil"
        folder4 = find_ton(m)
    elif re.match(r"^(SYSTX|TSTAT|EB-|33CS)", m):
        folder3 = "Controls"
    elif re.match(r"^(REME|F100F|F300E|HUM|CLEARSKY|ZEROH|SPLITHVAC|RPHI)", m):
        folder3 = "IAQ"

    return ParsedFolders(f1, f2, folder3, folder4, folder5)

def build_row5_desc(brand: str, model: str, scraped_desc: str, url: str, folders: ParsedFolders) -> str:
    parts: List[str] = []
    if scraped_desc:
        parts.append(scraped_desc)
    extra = " | ".join([x for x in [folders.folder3, folders.folder4, folders.folder5] if x])
    if extra:
        parts.append(extra)
    parts.append(f"Model: {model}")
    if url:
        parts.append(f"Source: {url}")
    parts.append("")
    parts.append(warranty_block(brand))
    return "\n".join(parts)


# ----------------------------
# Enricher
# ----------------------------

@dataclass
class ExtractedProduct:
    model: str
    name: str
    description: str
    source_url: str
    status: str

def score_url(url: str, model: str, domains: List[str]) -> int:
    u = (url or "").lower()
    m = model.lower()
    score = 0
    if any(d in u for d in domains):
        score += 4
    if m in u:
        score += 6
    if "product" in u or "products" in u:
        score += 2
    if "spec" in u or "specification" in u:
        score += 2
    if u.endswith(".pdf"):
        score += 3
    if "manual" in u:
        score -= 1
    if any(x in u for x in ["facebook", "pinterest", "reddit", "forums"]):
        score -= 4
    return score

class Enricher:
    def __init__(self, brand: str, catalog: ManufacturerCatalog, *, cache_path: str = "ingest_cache.sqlite",
                 rate_limit_search: float = 1.0, rate_limit_fetch: float = 0.8, learn_domains: bool = True):
        self.brand = brand.strip()
        self.catalog = catalog
        self.profile = self.catalog.get(self.brand)

        self.session = requests.Session()
        self.session.headers.update({"User-Agent": USER_AGENT})

        self.cache = SQLiteCache(cache_path)
        self.base_search = DuckDuckGoHTMLSearch(self.session, self.cache, rate_limit=rate_limit_search)
        self.fetcher = Fetcher(self.session, self.cache, rate_limit=rate_limit_fetch)
        self.learn_domains = learn_domains

    def close(self):
        self.cache.close()
        self.session.close()

    def _searcher(self) -> SearchProvider:
        return DomainPreferredSearch(self.base_search, self.profile.domains) if self.profile.domains else self.base_search

    def find_best_url(self, model: str) -> str:
        best_url = ""
        best_score = -999
        searcher = self._searcher()
        # try brand+model first if brand exists, then model alone
        queries = []
        if self.profile.name and self.profile.name.lower() != "unknown":
            queries.append(f"{self.profile.name} {model}")
        queries.append(model)

        for q in queries:
            for r in searcher.search(q, max_results=8):
                sc = score_url(r.url, model, self.profile.domains)
                if sc > best_score:
                    best_score, best_url = sc, r.url
            if best_score >= 9:
                break
        return best_url

    def extract_from_url(self, url: str, model: str) -> Tuple[str, str, str]:
        ct, payload = self.fetcher.get(url)
        for ex in EXTRACTORS:
            if ex.matches(url, self.brand):
                try:
                    name, desc = ex.extract(url=url, model=model, brand=self.brand, content_type=ct, payload=payload)
                    if name or desc:
                        return name, desc, ex.__class__.__name__
                except Exception:
                    continue
        return "", "", "None"

    def enrich_one(self, model: str) -> ExtractedProduct:
        model = norm_model(model)
        if not model:
            return ExtractedProduct(model="", name="", description="", source_url="", status="EMPTY_MODEL")

        ck = f"prod:{norm_brand(self.brand)}:{model}"
        cached = self.cache.get(ck)
        if cached:
            return ExtractedProduct(**json.loads(cached))

        url = self.find_best_url(model)
        if not url:
            out = ExtractedProduct(model=model, name="", description="", source_url="", status="NO_RESULT")
            self.cache.set(ck, json.dumps(dataclasses.asdict(out)))
            return out

        if self.learn_domains:
            self.catalog.learn(self.brand, url)
            # refresh profile in case we learned a good host
            self.profile = self.catalog.get(self.brand)

        name, desc, used = self.extract_from_url(url, model)
        status = "OK" if (name or desc) else "EXTRACT_EMPTY"
        out = ExtractedProduct(model=model, name=name, description=desc, source_url=url, status=f"{status}:{used}")
        self.cache.set(ck, json.dumps(dataclasses.asdict(out)))
        return out


# ----------------------------
# Workbook I/O
# ----------------------------

def col_to_idx(col: str) -> int:
    return column_index_from_string(col.upper())

def iter_models_from_sheet(ws: Worksheet, model_col: str, start_row: int) -> Iterable[Tuple[int, str]]:
    c = col_to_idx(model_col)
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=c).value
        m = norm_model(v)
        if m:
            yield r, m

def write_row(ws: Worksheet, row: int, *, folders: ParsedFolders, name: str, desc: str,
              overwrite: bool,
              folder_cols: Tuple[str,str,str,str,str]=("A","B","C","D","E"),
              name_col: str="V", desc_col: str="AE") -> None:
    folder_idx = [col_to_idx(c) for c in folder_cols]
    name_i = col_to_idx(name_col)
    desc_i = col_to_idx(desc_col)

    for idx, val in zip(folder_idx, [folders.folder1, folders.folder2, folders.folder3, folders.folder4, folders.folder5]):
        cur = ws.cell(row=row, column=idx).value
        if overwrite or (cur is None or str(cur).strip() == ""):
            ws.cell(row=row, column=idx).value = val

    cur = ws.cell(row=row, column=name_i).value
    if overwrite or (cur is None or str(cur).strip() == ""):
        if name:
            ws.cell(row=row, column=name_i).value = name

    cur = ws.cell(row=row, column=desc_i).value
    if overwrite or (cur is None or str(cur).strip() == ""):
        if desc:
            ws.cell(row=row, column=desc_i).value = desc


# ----------------------------
# CLI
# ----------------------------

def parse_args():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", help="Input XLSX to enrich")
    ap.add_argument("--sheet", default="Single part", help="Sheet name (default: Single part)")
    ap.add_argument("--brand", required=True, help="Brand/manufacturer name (e.g., Bryant)")
    ap.add_argument("--manufacturer-csv", default=None, help="CSV file: brand,domains (comma-separated)")
    ap.add_argument("--learned-domains", default="learned_domains.json", help="JSON file updated with learned hosts")
    ap.add_argument("--no-learn", action="store_true", help="Disable domain learning")
    ap.add_argument("--model-col", default="AD", help="Model column letter (default: AD)")
    ap.add_argument("--folder-cols", default="A,B,C,D,E", help="Comma-separated folder column letters (default: A,B,C,D,E). Use empty string for templates without folders.")
    ap.add_argument("--name-col", default="V", help="Part name column letter (default: V)")
    ap.add_argument("--desc-col", default="AE", help="Part description column letter (default: AE)")
    ap.add_argument("--bundle-mode", action="store_true", help="If set, also populate Product Name/Description using part fields when template is a bundle (1 part, 1 labor).")
    ap.add_argument("--product-name-col", default="F", help="Bundle Product Name column (default: F)")
    ap.add_argument("--product-desc-col", default="L", help="Bundle Product Description column (default: L)")
    ap.add_argument("--part-model-col", default=None, help="If provided, use this column for the part model number instead of --model-col (useful for bundle templates).")

    ap.add_argument("--start-row", type=int, default=4, help="Row to start scanning (default: 4)")
    ap.add_argument("--overwrite", action="store_true", help="Overwrite existing values")
    ap.add_argument("--out", required=True, help="Output XLSX path")
    ap.add_argument("--cache", default="ingest_cache.sqlite", help="SQLite cache path")
    ap.add_argument("--models", help="Text file with model numbers (one per line). Prints samples.")
    return ap.parse_args()


# ----------------------------
# Feedback memory (overrides + observations)
# ----------------------------

def fb_connect(path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(path)
    conn.execute(
        """CREATE TABLE IF NOT EXISTS overrides (
               brand TEXT NOT NULL,
               model TEXT NOT NULL,
               part_name TEXT,
               part_desc TEXT,
               folder1 TEXT,
               folder2 TEXT,
               folder3 TEXT,
               folder4 TEXT,
               folder5 TEXT,
               updated_at TEXT,
               PRIMARY KEY (brand, model)
           )"""
    )
    conn.execute(
        """CREATE TABLE IF NOT EXISTS observations (
               id INTEGER PRIMARY KEY AUTOINCREMENT,
               brand TEXT NOT NULL,
               model TEXT NOT NULL,
               source_url TEXT,
               part_name TEXT,
               part_desc TEXT,
               confidence REAL,
               issues TEXT,
               observed_at TEXT
           )"""
    )
    conn.commit()
    return conn

def fb_get_override(conn: sqlite3.Connection, brand: str, model: str) -> dict:
    cur = conn.execute(
        "SELECT part_name, part_desc, folder1, folder2, folder3, folder4, folder5 FROM overrides WHERE brand=? AND model=?",
        (brand.strip().upper(), model.strip().upper()),
    )
    row = cur.fetchone()
    if not row:
        return {}
    keys = ["part_name","part_desc","folder1","folder2","folder3","folder4","folder5"]
    return {k:v for k,v in zip(keys,row) if v not in (None,"")}

def fb_upsert_override(conn: sqlite3.Connection, brand: str, model: str, **fields) -> None:
    brand_u = brand.strip().upper()
    model_u = model.strip().upper()
    now = datetime.datetime.utcnow().isoformat(timespec="seconds") + "Z"
    cols = ["part_name","part_desc","folder1","folder2","folder3","folder4","folder5"]
    vals = [fields.get(c) for c in cols]
    conn.execute(
        """INSERT INTO overrides (brand, model, part_name, part_desc, folder1, folder2, folder3, folder4, folder5, updated_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
           ON CONFLICT(brand, model) DO UPDATE SET
             part_name=COALESCE(excluded.part_name, overrides.part_name),
             part_desc=COALESCE(excluded.part_desc, overrides.part_desc),
             folder1=COALESCE(excluded.folder1, overrides.folder1),
             folder2=COALESCE(excluded.folder2, overrides.folder2),
             folder3=COALESCE(excluded.folder3, overrides.folder3),
             folder4=COALESCE(excluded.folder4, overrides.folder4),
             folder5=COALESCE(excluded.folder5, overrides.folder5),
             updated_at=excluded.updated_at
        """,
        (brand_u, model_u, *vals, now),
    )
    conn.commit()

def fb_save_observation(conn: sqlite3.Connection, brand: str, model: str, source_url: str, part_name: str, part_desc: str, confidence: float, issues: str) -> None:
    now = datetime.datetime.utcnow().isoformat(timespec="seconds") + "Z"
    conn.execute(
        """INSERT INTO observations (brand, model, source_url, part_name, part_desc, confidence, issues, observed_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (brand.strip().upper(), model.strip().upper(), source_url or "", part_name or "", part_desc or "", float(confidence), issues or "", now),
    )
    conn.commit()

def compute_confidence(*, brand: str, model: str, name: str, desc: str, source_url: str) -> tuple[float, str]:
    issues = []
    score = 0.0
    n = (name or "").strip()
    d = (desc or "").strip()
    m = (model or "").strip().upper()

    if len(n) >= 10:
        score += 0.30
    else:
        issues.append("name_short_or_missing")

    if len(d) >= 200:
        score += 0.30
    else:
        issues.append("desc_short")

    if m and m in d.upper():
        score += 0.15
    else:
        issues.append("model_not_in_desc")

    # OEM domain hint
    host = ""
    try:
        from urllib.parse import urlparse
        host = (urlparse(source_url).netloc or "").lower()
    except Exception:
        host = ""
    brand_u = (brand or "").strip().upper()
    oem = False
    if brand_u == "BRYANT" and ("bryant.com" in host):
        oem = True
    if brand_u == "CARRIER" and ("carrier.com" in host):
        oem = True
    if brand_u in {"DAY & NIGHT","DAY AND NIGHT","DAYNIGHT","DAY&NIGHT"} and ("dayandnightcomfort.com" in host):
        oem = True
    if brand_u.startswith("MITSUBISHI") and ("mitsubishicomfort.com" in host):
        oem = True
    if oem:
        score += 0.15
    else:
        issues.append("non_oem_source")

    if re.search(r"\bwarranty\b|\byear\b|\blimited\b", d, re.I):
        score += 0.10
    else:
        issues.append("no_warranty_keywords")

    if d.lower().startswith("model:") or ("source:" in d.lower() and len(d) < 120):
        score -= 0.15
        issues.append("looks_placeholder")

    score = max(0.0, min(1.0, score))
    return score, ",".join(issues)

def ensure_review_sheet(wb, name: str):
    if name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row-1)
        return ws
    ws = wb.create_sheet(name)
    ws.append([
        "Row","Brand","Model","Confidence","Issues","Source URL",
        "Current Part Name","Current Part Description",
        "Override Part Name","Override Part Description",
        "Override Folder1","Override Folder2","Override Folder3","Override Folder4","Override Folder5",
        "Notes"
    ])
    return ws


def write_review_row(ws, model: str, status: str, confidence: float | None, source: str | None):
    """Append a minimal review row."""
    r = ws.max_row + 1
    ws.cell(row=r, column=1, value=model)
    ws.cell(row=r, column=2, value=status)
    ws.cell(row=r, column=3, value=float(confidence) if confidence is not None else None)
    ws.cell(row=r, column=4, value=source or '')
def main() -> int:
    args = parse_args()
    warranty_rules = load_rules_json(args.warranty_rules)
    series_rules = load_rules_json(args.series_rules)
    catalog = ManufacturerCatalog(args.manufacturer_csv, learned_path=args.learned_domains)
    enricher = Enricher(args.brand, catalog, cache_path=args.cache, learn_domains=(not args.no_learn))

    try:
        if args.models:
            models = [norm_model(x) for x in open(args.models, "r", encoding="utf-8").read().splitlines()]
            models = [m for m in models if m]
            print(f"Loaded {len(models)} models")
            for m in models[:30]:
                prod = enricher.enrich_one(m)
                folders = parse_folders(args.brand, m)
                desc = build_row5_desc(args.brand, m, prod.description, prod.source_url, folders)
                print("\n---")
                print("Model:", m, "Status:", prod.status)
                print("Name:", prod.name)
                print(desc[:320] + ("..." if len(desc) > 320 else ""))
            return 0

        if not args.input:
            raise SystemExit("Provide --input XLSX or --models text file.")

        wb = load_workbook(args.input)
        fb_conn = fb_connect(args.feedback_db) if args.feedback_db else None
        if args.sheet not in wb.sheetnames:
            raise SystemExit(f"Sheet '{args.sheet}' not found. Available: {wb.sheetnames}")
        ws = wb[args.sheet]
        review_ws = ensure_review_sheet(wb, args.review_sheet) if args.write_review else None

        processed = ok = miss = 0
        model_col_to_use = args.part_model_col or args.model_col
        for row, model in iter_models_from_sheet(ws, model_col_to_use, args.start_row):
            folders = parse_folders(args.brand, model)
            prod = enricher.enrich_one(model)

            # Series/name rule fallback (only when scraping didn't return a strong name)
            sr = first_rule_match(series_rules, args.brand, model)
            if sr and (not prod.name or len((prod.name or '').strip()) < 8):
                prefix = (sr.get('name_prefix') or args.brand).strip()
                series = (sr.get('series') or '').strip()
                etype = (sr.get('type') or '').strip()
                bits = [b for b in [prefix, series, etype, model] if b]
                prod.name = ' '.join(bits)

            desc = build_row5_desc(args.brand, model, prod.description, prod.source_url, folders)
            write_row(ws, row, folders=folders, name=prod.name, desc=desc, overwrite=args.overwrite)

            if review_ws is not None:
                # Write a lightweight review row: status + confidence + source
                write_review_row(review_ws, model=model, status=prod.status, confidence=prod.confidence, source=prod.source_url)

            processed += 1
            if (prod.status or '').upper().startswith('OK'):
                ok += 1
            else:
                miss += 1

            if processed % 25 == 0:
                print(f"Processed {processed} rows | OK={ok} Miss={miss}", flush=True)

        wb.save(args.out)
        if fb_conn is not None:
            try:
                fb_conn.close()
            except Exception:
                pass
        print(f"Saved: {args.out}")
        print(f"Processed={processed} OK={ok} Miss={miss}")
        print(f"Learned domains saved to: {args.learned_domains}")
        return 0
    finally:
        enricher.close()

if __name__ == "__main__":
    sys.exit(main())

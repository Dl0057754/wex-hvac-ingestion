#!/usr/bin/env python3
"""
convert_pricebook_to_wex_skeleton.py

Stage A converter:
- Ingest distributor "pricebooks" in many formats (xlsx/xlsm/csv/pdf/docx/txt + optional images via OCR).
- Extract model/part numbers and (optionally) pricing.
- Write a WEX FSM "single part" upload skeleton workbook where Column AD contains Model/Part Number.
- Preserves a best-effort "category context" from merged heading rows / section headers.
- Produces a companion CSV report of extracted items and where they came from.

This script is intentionally conservative:
- It NEVER stops on a bad row/page; it logs and continues.
- It separates concerns: conversion builds a skeleton; enrichment fills folders/names/descriptions.

Dependencies:
  pip install openpyxl pandas python-docx requests
Optional (recommended):
  pip install pdfplumber beautifulsoup4 lxml
Optional OCR (only if you truly need images):
  pip install pytesseract pillow
  and install Tesseract binary separately (system-level).

Usage examples:

1) Convert an XLSX pricebook into a skeleton:
  python convert_pricebook_to_wex_skeleton.py \
    --input "Distributor_Bryant.xlsx" \
    --template "Single part template.xlsx" \
    --brand "Bryant" \
    --out "WEX_skeleton_Bryant.xlsx"

2) Convert a whole folder (mix of files):
  python convert_pricebook_to_wex_skeleton.py \
    --input "pricebooks_in/" \
    --template "Single part template.xlsx" \
    --brand "Carrier" \
    --out "WEX_skeleton_Carrier.xlsx"

3) Use a mapping file for tricky distributors:
  python convert_pricebook_to_wex_skeleton.py \
    --input "Distributor_X.xlsx" \
    --template "Single part template.xlsx" \
    --brand "Bryant" \
    --mapping "mapping.json" \
    --out "WEX_skeleton.xlsx"

Mapping format (optional) - lets you pin down sheet + columns precisely:
{
  "files": [
    {
      "match": "Distributor_X",
      "sheets": [
        {"name": "Furnaces", "model_cols": ["B"], "price_cols": ["H"], "skip_rows": 2},
        {"name": "Outdoor",  "model_cols": ["C"], "price_cols": ["J"]}
      ]
    }
  ]
}

If mapping is absent, the script auto-detects model numbers anywhere in the sheet.

Output:
- XLSX skeleton (your template populated with extracted rows)
- CSV report: <out>.extracted.csv (same base name)

Notes:
- Column AD is assumed to be "Part Model Number" per your WEX template convention.
- If your template uses a different sheet name than "Single part", pass --template-sheet.
"""

from __future__ import annotations

import argparse
import csv
import dataclasses
import json
import os
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

try:
    import pdfplumber  # type: ignore
except Exception:
    pdfplumber = None  # type: ignore

try:
    import docx  # python-docx
except Exception:
    docx = None  # type: ignore

try:
    import pytesseract  # type: ignore
    from PIL import Image  # type: ignore
except Exception:
    pytesseract = None  # type: ignore
    Image = None  # type: ignore


# ----------------------------
# Heuristics
# ----------------------------

# A model/part token: letters+digits, may include hyphen, ends with letters/digits.
# Tuned to your HVAC patterns (Carrier/Bryant etc.) but works broadly.
# A model/part token: HVAC is messy. We use multiple patterns and then normalize.
# - Plain tokens: 191VAN02400W
# - Hyphenated tokens: FF-2401C05, 331831-701
# - Mixed: HUMCRSBP2412-A18
MODEL_RE_LIST = [
    re.compile(r"\b[A-Z0-9][A-Z0-9\-]{6,36}[A-Z0-9]\b", re.I),
    re.compile(r"\b[A-Z0-9]{2,10}\-[A-Z0-9\-]{2,30}\b", re.I),
]
# For extraction from surrounding text like "Model: XXXXX"
MODEL_LABEL_RE = re.compile(r"(?:MODEL|MOD)\s*[:#]?\s*([A-Z0-9][A-Z0-9\-]{5,36})", re.I)


# Price patterns: $1,234.56 or 1234.56 or 1,234
PRICE_RE = re.compile(r"(?:\$\s*)?(\d{1,3}(?:,\d{3})+|\d+)(?:\.(\d{2}))?")

# Headings: if a row has mostly text and it's merged across many columns, treat as section
def looks_like_heading(text: str) -> bool:
    t = (text or "").strip()
    if not t:
        return False
    if len(t) < 4:
        return False
    if any(x in t.lower() for x in ["page", "price", "msrp", "list", "notes"]):
        return True
    # mostly letters/spaces
    letters = sum(ch.isalpha() for ch in t)
    digits = sum(ch.isdigit() for ch in t)
    return letters >= 6 and digits <= 2


def norm(s: str) -> str:
    return re.sub(r"\s+", "", (s or "").strip()).upper()


def col_to_idx(col: str) -> int:
    return column_index_from_string(col.upper())


# ----------------------------
# Data model
# ----------------------------

@dataclass
class ExtractedItem:
    brand: str
    model: str
    price: str
    category_1: str
    category_2: str
    source_file: str
    source_sheet: str
    source_ref: str  # row/col or page/para
    raw_context: str


# ----------------------------
# Mapping support
# ----------------------------

def load_mapping(path: Optional[str]) -> Dict:
    if not path:
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def match_file_mapping(mapping: Dict, filename: str) -> Optional[Dict]:
    files = mapping.get("files") or []
    for f in files:
        m = (f.get("match") or "").strip()
        if m and m.lower() in filename.lower():
            return f
    return None


# ----------------------------
# Readers
# ----------------------------

def extract_from_xlsx(path: Path, brand: str, mapping: Optional[Dict]) -> List[ExtractedItem]:
    items: List[ExtractedItem] = []
    wb = load_workbook(path, data_only=True)

    file_map = match_file_mapping(mapping or {}, path.name) if mapping else None

    # Determine which sheets to parse
    if file_map and file_map.get("sheets"):
        sheet_specs = file_map["sheets"]
    else:
        sheet_specs = [{"name": s} for s in wb.sheetnames]

    for spec in sheet_specs:
        sname = spec.get("name")
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]

        # Pre-compute merged ranges for quick lookup
        merged_ranges = list(ws.merged_cells.ranges)

        def is_merged_row(r: int) -> bool:
            for mr in merged_ranges:
                if mr.min_row <= r <= mr.max_row and (mr.max_col - mr.min_col) >= 3:
                    return True
            return False

        # Optional explicit columns
        model_cols = spec.get("model_cols") or []
        price_cols = spec.get("price_cols") or []
        skip_rows = int(spec.get("skip_rows") or 0)

        # Heading stack, updated as we move down
        current_heading_1 = ""
        current_heading_2 = ""

        max_r = ws.max_row
        max_c = min(ws.max_column, 80)  # don't scan 500 empty columns

        for r in range(1 + skip_rows, max_r + 1):
            # detect heading rows
            row_values = [ws.cell(row=r, column=c).value for c in range(1, max_c + 1)]
            row_texts = [str(v).strip() for v in row_values if v is not None and str(v).strip()]

            if row_texts and is_merged_row(r):
                # Pick the longest text cell as heading candidate
                cand = max(row_texts, key=len)
                if looks_like_heading(cand):
                    # shift heading stack
                    current_heading_1, current_heading_2 = current_heading_2 or current_heading_1, cand.strip()

            # Extract models + row-level price
            found_models: List[Tuple[str, str]] = []
            found_prices: Dict[str, str] = {}  # model -> price (we use _row)

            def scan_cell_for_model_and_price(v, cidx: int) -> None:
                if v is None:
                    return
                txt_cell = str(v).strip()
                if not txt_cell:
                    return

                # find models (multiple patterns + label-based)
                candidates: List[str] = []
                ml = MODEL_LABEL_RE.search(txt_cell.upper())
                if ml:
                    candidates.append(ml.group(1))

                for rx in MODEL_RE_LIST:
                    candidates.extend(rx.findall(txt_cell.upper()))

                for m in candidates:
                    mm = norm(m)
                    # reject obvious non-model junk (pure digits, too short)
                    if mm.isdigit() or len(mm) < 6:
                        continue
                    ctx = txt_cell[:160]
                    found_models.append((mm, ctx))
                # find prices
                pm = PRICE_RE.search(txt.replace(" ", ""))
                if pm:
                    dollars = pm.group(1).replace(",", "")
                    cents = pm.group(2) or ""
                    price = dollars + (("." + cents) if cents else "")
                    # store as generic row price; we'll attach to first model in row if needed
                    found_prices["_row"] = price

            if model_cols:
                for col in model_cols:
                    c = col_to_idx(col)
                    scan_cell_for_model_and_price(ws.cell(row=r, column=c).value, c)
            else:
                # auto-scan across the row
                for c in range(1, max_c + 1):
                    scan_cell_for_model_and_price(ws.cell(row=r, column=c).value, c)

            # explicit prices if provided
            if price_cols:
                for col in price_cols:
                    c = col_to_idx(col)
                    v = ws.cell(row=r, column=c).value
                    if v is None:
                        continue
                    txt = str(v).strip()
                    pm = PRICE_RE.search(txt.replace(" ", ""))
                    if pm:
                        dollars = pm.group(1).replace(",", "")
                        cents = pm.group(2) or ""
                        found_prices["_row"] = dollars + (("." + cents) if cents else "")

            if not found_models:
                continue

            # De-dupe models in the row
            seen = set()
            uniq_models = []
            for m, ctx in found_models:
                if m in seen:
                    continue
                seen.add(m)
                uniq_models.append((m, ctx))

            # Attach price (row-level)
            row_price = found_prices.get("_row", "")

            for m, ctx in uniq_models:
                items.append(
                    ExtractedItem(
                        brand=brand,
                        model=m,
                        price=row_price,
                        category_1=current_heading_1.strip(),
                        category_2=current_heading_2.strip(),
                        source_file=path.name,
                        source_sheet=sname,
                        source_ref=f"R{r}",
                        raw_context=ctx,
                    )
                )

    return items


def extract_from_csv(path: Path, brand: str) -> List[ExtractedItem]:
    items: List[ExtractedItem] = []
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    for i, row in df.iterrows():
        joined = " | ".join([str(x) for x in row.values if str(x).strip()])
        models = [norm(m) for m in MODEL_RE.findall(joined.upper())]
        models = [m for m in dict.fromkeys(models) if m and not m.isdigit()]
        if not models:
            continue
        pm = PRICE_RE.search(joined.replace(" ", ""))
        price = ""
        if pm:
            dollars = pm.group(1).replace(",", "")
            cents = pm.group(2) or ""
            price = dollars + (("." + cents) if cents else "")
        for m in models:
            items.append(
                ExtractedItem(
                    brand=brand,
                    model=m,
                    price=price,
                    category_1="",
                    category_2="",
                    source_file=path.name,
                    source_sheet="(csv)",
                    source_ref=f"row:{i+1}",
                    raw_context=joined[:160],
                )
            )
    return items


def extract_from_pdf(path: Path, brand: str) -> List[ExtractedItem]:
    items: List[ExtractedItem] = []
    if not pdfplumber:
        return items
    with pdfplumber.open(str(path)) as pdf:  # type: ignore
        for pageno, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            if not text.strip():
                continue
            lines = text.splitlines()
            for li, line in enumerate(lines, start=1):
                models = [norm(m) for m in MODEL_RE.findall(line.upper())]
                models = [m for m in dict.fromkeys(models) if m and not m.isdigit()]
                if not models:
                    continue
                pm = PRICE_RE.search(line.replace(" ", ""))
                price = ""
                if pm:
                    dollars = pm.group(1).replace(",", "")
                    cents = pm.group(2) or ""
                    price = dollars + (("." + cents) if cents else "")
                for m in models:
                    items.append(
                        ExtractedItem(
                            brand=brand,
                            model=m,
                            price=price,
                            category_1="",
                            category_2="",
                            source_file=path.name,
                            source_sheet="(pdf)",
                            source_ref=f"page:{pageno} line:{li}",
                            raw_context=line.strip()[:160],
                        )
                    )
    return items


def extract_from_docx(path: Path, brand: str) -> List[ExtractedItem]:
    items: List[ExtractedItem] = []
    if not docx:
        return items
    doc = docx.Document(str(path))  # type: ignore
    paras = [p.text for p in doc.paragraphs if p.text and p.text.strip()]
    for i, para in enumerate(paras, start=1):
        models = [norm(m) for m in MODEL_RE.findall(para.upper())]
        models = [m for m in dict.fromkeys(models) if m and not m.isdigit()]
        if not models:
            continue
        pm = PRICE_RE.search(para.replace(" ", ""))
        price = ""
        if pm:
            dollars = pm.group(1).replace(",", "")
            cents = pm.group(2) or ""
            price = dollars + (("." + cents) if cents else "")
        for m in models:
            items.append(
                ExtractedItem(
                    brand=brand,
                    model=m,
                    price=price,
                    category_1="",
                    category_2="",
                    source_file=path.name,
                    source_sheet="(docx)",
                    source_ref=f"para:{i}",
                    raw_context=para.strip()[:160],
                )
            )
    return items


def extract_from_image(path: Path, brand: str) -> List[ExtractedItem]:
    items: List[ExtractedItem] = []
    if not pytesseract or not Image:
        return items
    try:
        img = Image.open(str(path))  # type: ignore
        text = pytesseract.image_to_string(img)  # type: ignore
    except Exception:
        return items
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    for i, line in enumerate(lines, start=1):
        models = [norm(m) for m in MODEL_RE.findall(line.upper())]
        models = [m for m in dict.fromkeys(models) if m and not m.isdigit()]
        if not models:
            continue
        pm = PRICE_RE.search(line.replace(" ", ""))
        price = ""
        if pm:
            dollars = pm.group(1).replace(",", "")
            cents = pm.group(2) or ""
            price = dollars + (("." + cents) if cents else "")
        for m in models:
            items.append(
                ExtractedItem(
                    brand=brand,
                    model=m,
                    price=price,
                    category_1="",
                    category_2="",
                    source_file=path.name,
                    source_sheet="(image)",
                    source_ref=f"line:{i}",
                    raw_context=line[:160],
                )
            )
    return items


# ----------------------------
# Skeleton writer
# ----------------------------

def write_skeleton(template_path: Path, profile: Dict, out_path: Path, items: List[ExtractedItem]) -> None:
    wb = load_workbook(str(template_path))
    template_sheet = profile.get('template_sheet')
    if template_sheet not in wb.sheetnames:
        raise SystemExit(f"Template sheet '{template_sheet}' not found. Available: {wb.sheetnames}")
    ws = wb[template_sheet]

    # Find first empty row by checking column AD
    model_col_idx = col_to_idx(profile.get("model_col","AD"))
    start_row = int(profile.get("start_row", 4))
    r = start_row
    while r <= ws.max_row and (ws.cell(row=r, column=model_col_idx).value not in (None, "")):
        r += 1

    # Columns assumed by your template convention
    folder_cols = profile.get("folder_cols", ["A","B","C","D","E"])
    part_name_col = profile.get("part_name_col", "V")
    part_desc_col = profile.get("part_desc_col", "AE")

    def set_cell(row: int, col_letter: str, value: str):
        c = col_to_idx(col_letter)
        if value is None:
            return
        ws.cell(row=row, column=c).value = value

    # Write rows
    for it in items:
        # Minimal skeleton: put model in AD, brand in Folder2, headings as optional placeholders
        set_cell(r, profile.get("model_col","AD"), it.model)
        # Folder columns are optional (some templates do not have them)
        if folder_cols:
            if len(folder_cols) >= 1:
                set_cell(r, folder_cols[0], "Equipment")
            if len(folder_cols) >= 2:
                set_cell(r, folder_cols[1], it.brand)
            # best-effort category placeholders
            if len(folder_cols) >= 3:
                if it.category_2:
                    set_cell(r, folder_cols[2], it.category_2)
                elif it.category_1:
                    set_cell(r, folder_cols[2], it.category_1)
            if len(folder_cols) >= 4:
                if it.category_1 and it.category_2 and it.category_1 != it.category_2:
                    set_cell(r, folder_cols[3], it.category_1)

        # Put price into a note-like area if your template has a place; otherwise keep in description placeholder
        if it.price:
            set_cell(r, part_desc_col, f"Distributor price: {it.price}")

        # Keep a source breadcrumb in description placeholder too (safe to overwrite later)
        breadcrumb = f"Source: {it.source_file} | {it.source_sheet} | {it.source_ref}"
        cur = ws.cell(row=r, column=col_to_idx(part_desc_col)).value
        if cur:
            set_cell(r, part_desc_col, str(cur) + "\n" + breadcrumb)
        else:
            set_cell(r, part_desc_col, breadcrumb)

        r += 1

    wb.save(str(out_path))


def write_report_csv(out_xlsx: Path, items: List[ExtractedItem]) -> Path:
    report_path = out_xlsx.with_suffix(out_xlsx.suffix + ".extracted.csv")
    with open(report_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(dataclasses.asdict(items[0]).keys() if items else [
            "brand","model","price","category_1","category_2","source_file","source_sheet","source_ref","raw_context"
        ])
        for it in items:
            w.writerow([
                it.brand, it.model, it.price, it.category_1, it.category_2,
                it.source_file, it.source_sheet, it.source_ref, it.raw_context
            ])
    return report_path



# ----------------------------
# Template Profiles
# ----------------------------

def load_templates_config(path: Optional[str]) -> Dict:
    if not path:
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def pick_profile(templates_cfg: Dict, profile_name: Optional[str]) -> Dict:
    if not templates_cfg:
        return {}
    if profile_name and profile_name in templates_cfg:
        return templates_cfg[profile_name]
    # fallback: if only one profile exists, use it
    if len(templates_cfg) == 1:
        return next(iter(templates_cfg.values()))
    return {}

def detect_template_sheet_and_headers(template_path: Path) -> Tuple[str, int, Dict[str, str]]:
    """
    Best-effort auto-detection:
    - Finds the first sheet containing recognizable headers like "Part Model Number" / "Manufacturer #" / "Part Model Number".
    - Returns: (sheet_name, header_row, header_map{normalized_header -> column_letter})
    """
    wb = load_workbook(str(template_path), data_only=True)
    header_candidates = [
        "part model number", "model number", "manufacturer #", "manufacturer", "part name", "description",
        "folder 1", "folder1", "part_description", "part description", "part_Description".lower()
    ]

    def normalize_header(x: str) -> str:
        return re.sub(r"\s+", " ", (x or "").strip().lower())

    from openpyxl.utils import get_column_letter

    for s in wb.sheetnames:
        ws = wb[s]
        max_r = min(ws.max_row, 30)
        max_c = min(ws.max_column, 80)
        for r in range(1, max_r + 1):
            row = []
            for c in range(1, max_c + 1):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                txt = normalize_header(str(v))
                if not txt:
                    continue
                row.append((txt, c))
            if not row:
                continue
            hit = sum(1 for txt, _ in row if any(k in txt for k in header_candidates))
            if hit >= 3:  # looks like a header row
                header_map = {txt: get_column_letter(c) for txt, c in row}
                return s, r, header_map
    # fallback
    return wb.sheetnames[0], 1, {}

def resolve_template_profile(template_path: Path, templates_cfg: Dict, profile_name: Optional[str], template_sheet_override: Optional[str]) -> Dict:
    prof = pick_profile(templates_cfg, profile_name)
    if prof:
        prof = dict(prof)  # copy
        if template_sheet_override:
            prof["template_sheet"] = template_sheet_override
        return prof

    # auto-detect
    sheet, header_row, header_map = detect_template_sheet_and_headers(template_path)

    def find_col(*needles: str) -> str:
        for n in needles:
            n2 = re.sub(r"\s+", " ", n.strip().lower())
            for h, col in header_map.items():
                if n2 == h or n2 in h:
                    return col
        return ""

    folder_cols = []
    # prefer explicit Folder 1..5
    for i in range(1,6):
        c = find_col(f"folder {i}")
        if c:
            folder_cols.append(c)

    model_col = find_col("part model number", "manufacturer #", "model number")
    part_name_col = find_col("part name", "product name")
    desc_col = find_col("part_description", "part description", "description")

    start_row = header_row + 2  # leave a buffer row for notes/requirements
    prof = {
        "template_sheet": template_sheet_override or sheet,
        "start_row": start_row,
        "model_col": model_col or "AD",
        "folder_cols": folder_cols,
        "part_name_col": part_name_col or "V",
        "part_desc_col": desc_col or "AE",
        "_detected_header_row": header_row,
    }
    return prof

# ----------------------------
# Main
# ----------------------------

def gather_inputs(input_path: str) -> List[Path]:
    p = Path(input_path)
    if p.is_dir():
        files = []
        for ext in ["*.xlsx","*.xlsm","*.csv","*.pdf","*.docx","*.txt","*.png","*.jpg","*.jpeg"]:
            files.extend(p.glob(ext))
        return sorted(files)
    if p.exists():
        return [p]
    raise SystemExit(f"Input path not found: {input_path}")


def extract_from_textfile(path: Path, brand: str) -> List[ExtractedItem]:
    items: List[ExtractedItem] = []
    txt = path.read_text(encoding="utf-8", errors="ignore")
    lines = [ln.strip() for ln in txt.splitlines() if ln.strip()]
    for i, line in enumerate(lines, start=1):
        models = [norm(m) for m in MODEL_RE.findall(line.upper())]
        models = [m for m in dict.fromkeys(models) if m and not m.isdigit()]
        if not models:
            continue
        pm = PRICE_RE.search(line.replace(" ", ""))
        price = ""
        if pm:
            dollars = pm.group(1).replace(",", "")
            cents = pm.group(2) or ""
            price = dollars + (("." + cents) if cents else "")
        for m in models:
            items.append(ExtractedItem(
                brand=brand, model=m, price=price,
                category_1="", category_2="",
                source_file=path.name, source_sheet="(txt)", source_ref=f"line:{i}",
                raw_context=line[:160],
            ))
    return items


def parse_args():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Input file or folder of pricebooks")
    ap.add_argument("--template", required=True, help="WEX upload template XLSX (single part template)")
    ap.add_argument("--template-sheet", default=None, help="Template sheet name (overrides profile). If omitted, uses profile or auto-detect.")
    ap.add_argument("--templates-config", default=None, help="JSON config of template profiles (e.g., wex_templates.json)")
    ap.add_argument("--template-profile", default=None, help="Template profile name in templates-config (e.g., wex_single_part, wex_supplier_loader, wex_bundle_1part_1labor)")
    ap.add_argument("--brand", required=True, help="Brand name to tag output rows (e.g., Bryant)")
    ap.add_argument("--mapping", default=None, help="Optional JSON mapping for known distributor layouts")
    ap.add_argument("--out", required=True, help="Output skeleton XLSX")
    ap.add_argument("--dedupe", action="store_true", help="De-dupe models across all inputs")
    return ap.parse_args()


def main() -> int:
    args = parse_args()
    mapping = load_mapping(args.mapping)
    inputs = gather_inputs(args.input)
    brand = args.brand.strip()

    all_items: List[ExtractedItem] = []
    for p in inputs:
        ext = p.suffix.lower()
        try:
            if ext in [".xlsx", ".xlsm"]:
                all_items.extend(extract_from_xlsx(p, brand, mapping))
            elif ext == ".csv":
                all_items.extend(extract_from_csv(p, brand))
            elif ext == ".pdf":
                all_items.extend(extract_from_pdf(p, brand))
            elif ext == ".docx":
                all_items.extend(extract_from_docx(p, brand))
            elif ext in [".png", ".jpg", ".jpeg"]:
                all_items.extend(extract_from_image(p, brand))
            elif ext == ".txt":
                all_items.extend(extract_from_textfile(p, brand))
        except Exception as e:
            print(f"[WARN] Failed to parse {p.name}: {e}", file=sys.stderr)
            continue

    if args.dedupe:
        seen = set()
        deduped = []
        for it in all_items:
            k = (it.brand.upper(), it.model.upper())
            if k in seen:
                continue
            seen.add(k)
            deduped.append(it)
        all_items = deduped

    if not all_items:
        raise SystemExit("No models extracted. If this is an image/PDF, install pdfplumber or OCR dependencies.")

    out_xlsx = Path(args.out)
    templates_cfg = load_templates_config(args.templates_config)
    profile = resolve_template_profile(Path(args.template), templates_cfg, args.template_profile, args.template_sheet)
    write_skeleton(Path(args.template), profile, out_xlsx, all_items)
    report_path = write_report_csv(out_xlsx, all_items)

    print(f"Extracted items: {len(all_items)}")
    print(f"Saved skeleton: {out_xlsx}")
    print(f"Saved report:   {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
